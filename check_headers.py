from openpyxl import load_workbook
p='D:/armtemp/ADF_Analyzer_v10_Production/output/adf_analysis_latest.xlsx'
wb=load_workbook(p, read_only=True)
ws=wb['Activities']
headers=[cell.value for cell in next(ws.iter_rows(min_row=1,max_row=1))]
print('Headers:', headers)
print('Contains ExecutionStage exact:', 'ExecutionStage' in (headers or []))
print('Lowercase match:', any(h and h.lower().startswith('execution') for h in headers))
ws2=wb['ActivityExecutionOrder']
headers2=[cell.value for cell in next(ws2.iter_rows(min_row=1,max_row=1))]
print('ActivityExecutionOrder headers:', headers2)
print('Contains FromExecutionStage:', 'FromExecutionStage' in (headers2 or []))
print('Contains ToExecutionStage:', 'ToExecutionStage' in (headers2 or []))
